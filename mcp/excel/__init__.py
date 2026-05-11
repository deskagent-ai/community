#!/usr/bin/env python3
# Copyright (C) 2026 realvirtual GmbH
# This file is part of DeskAgent.
# AGPL-3.0-or-later - see LICENSE for details.


"""
Excel MCP Server
================
MCP Server für das Lesen und Schreiben von Excel-Dateien (.xlsx).
Unterstützt Zellen, Bereiche und Arbeitsblätter.
"""

import json
from pathlib import Path
from typing import Any, List, Dict

from mcp.server.fastmcp import FastMCP

# Excel library
try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.utils import get_column_letter, column_index_from_string
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# DeskAgent MCP API (provides config, paths, logging via HTTP)
from _mcp_api import load_config, get_workspace_dir

mcp = FastMCP("excel")

# Tool metadata for dynamic icon/color in WebUI
TOOL_METADATA = {
    "icon": "table_chart",
    "color": "#217346"  # Excel green
}

# Integration schema for Settings UI
INTEGRATION_SCHEMA = {
    "name": "Excel",
    "icon": "table_chart",
    "color": "#217346",
    "config_key": None,  # Keine Config noetig
    "auth_type": "none",
}

# Tools that return external/untrusted content (prompt injection risk)
# These will be wrapped with sanitization by the anonymization proxy
HIGH_RISK_TOOLS = {
    "excel_read_sheet",
    "excel_read_cell",
    "excel_read_range",
}

# Read-only tools that only retrieve data (for tool_mode: "read_only")
READ_ONLY_TOOLS = {
    "excel_get_info",
    "excel_list_sheets",
    "excel_read_cell",
    "excel_read_range",
    "excel_read_sheet",
}

# Tools that modify data
DESTRUCTIVE_TOOLS = {
    "excel_write_cell",
    "excel_write_range",
    "excel_write_sheet",
    "excel_create_sheet",
    "excel_delete_sheet",
}


def is_configured() -> bool:
    """Prüft ob Excel MCP verfügbar ist.

    Excel MCP ist verfügbar wenn openpyxl installiert ist.
    Kann über excel.enabled deaktiviert werden.
    """
    if not OPENPYXL_AVAILABLE:
        return False

    config = load_config()
    mcp_config = config.get("excel", {})

    if mcp_config.get("enabled") is False:
        return False

    return True


def _resolve_path(file_path: str) -> Path:
    """Resolve file path (absolute or relative to workspace)."""
    path = Path(file_path)
    if not path.is_absolute():
        workspace = get_workspace_dir()
        path = workspace / path
    return path


def _cell_to_tuple(cell_ref: str) -> tuple:
    """Convert cell reference like 'A1' to (row, col) tuple (1-indexed)."""
    import re
    match = re.match(r'^([A-Z]+)(\d+)$', cell_ref.upper())
    if not match:
        raise ValueError(f"Invalid cell reference: {cell_ref}")
    col_letter, row_num = match.groups()
    col = column_index_from_string(col_letter)
    row = int(row_num)
    return (row, col)


def _parse_range(range_ref: str) -> tuple:
    """Parse range like 'A1:B10' to ((start_row, start_col), (end_row, end_col))."""
    if ':' not in range_ref:
        raise ValueError(f"Invalid range reference: {range_ref}")
    start_cell, end_cell = range_ref.split(':')
    start = _cell_to_tuple(start_cell)
    end = _cell_to_tuple(end_cell)
    return (start, end)


@mcp.tool()
def excel_get_info(file_path: str) -> str:
    """
    Zeigt Informationen über eine Excel-Datei.

    Listet alle Arbeitsblätter und deren Größe auf.

    Args:
        file_path: Pfad zur Excel-Datei (.xlsx)

    Returns:
        JSON mit Datei-Informationen
    """
    try:
        path = _resolve_path(file_path)
        if not path.exists():
            return f"Fehler: Datei nicht gefunden: {path}"

        wb = load_workbook(path, read_only=True, data_only=True)

        sheets_info = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            max_row = ws.max_row
            max_col = ws.max_column
            sheets_info.append({
                "name": sheet_name,
                "rows": max_row,
                "columns": max_col,
                "size": f"{get_column_letter(max_col)}{max_row}"
            })

        wb.close()

        result = {
            "file": str(path),
            "sheet_count": len(sheets_info),
            "sheets": sheets_info
        }

        return json.dumps(result, ensure_ascii=False, indent=2)

    except Exception as e:
        return f"Fehler: {str(e)}"


@mcp.tool()
def excel_list_sheets(file_path: str) -> str:
    """
    Listet alle Arbeitsblätter einer Excel-Datei auf.

    Args:
        file_path: Pfad zur Excel-Datei (.xlsx)

    Returns:
        JSON-Array mit Arbeitsblattnamen
    """
    try:
        path = _resolve_path(file_path)
        if not path.exists():
            return f"Fehler: Datei nicht gefunden: {path}"

        wb = load_workbook(path, read_only=True)
        sheets = wb.sheetnames
        wb.close()

        return json.dumps(sheets, ensure_ascii=False)

    except Exception as e:
        return f"Fehler: {str(e)}"


@mcp.tool()
def excel_read_cell(file_path: str, cell_ref: str, sheet_name: str = None, preserve_formulas: bool = False) -> str:
    """
    Liest den Wert einer einzelnen Zelle.

    Args:
        file_path: Pfad zur Excel-Datei (.xlsx)
        cell_ref: Zellenreferenz (z.B. 'A1', 'B5')
        sheet_name: Name des Arbeitsblatts (optional, Standard: erstes Blatt)
        preserve_formulas: Formeln als String lesen statt berechnete Werte (Standard: false)

    Returns:
        Zellenwert als String (oder Formel wenn preserve_formulas=true)
    """
    try:
        path = _resolve_path(file_path)
        if not path.exists():
            return f"Fehler: Datei nicht gefunden: {path}"

        wb = load_workbook(path, read_only=True, data_only=not preserve_formulas)

        # Get worksheet
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                wb.close()
                return f"Fehler: Arbeitsblatt '{sheet_name}' nicht gefunden"
            ws = wb[sheet_name]
        else:
            ws = wb.active

        # Read cell
        cell = ws[cell_ref.upper()]
        value = cell.value

        wb.close()

        # Convert to string
        if value is None:
            return ""
        else:
            return str(value)

    except Exception as e:
        return f"Fehler: {str(e)}"


@mcp.tool()
def excel_read_range(file_path: str, range_ref: str, sheet_name: str = None, include_header: bool = True, preserve_formulas: bool = False) -> str:
    """
    Liest einen Zellenbereich und gibt ihn als JSON zurück.

    Args:
        file_path: Pfad zur Excel-Datei (.xlsx)
        range_ref: Bereichsreferenz (z.B. 'A1:C10')
        sheet_name: Name des Arbeitsblatts (optional, Standard: erstes Blatt)
        include_header: Erste Zeile als Spaltennamen verwenden (Standard: true)
        preserve_formulas: Formeln als String lesen statt berechnete Werte (Standard: false)

    Returns:
        JSON-Array mit Zeilen (als Dictionaries wenn include_header=true, sonst als Arrays)
    """
    try:
        path = _resolve_path(file_path)
        if not path.exists():
            return f"Fehler: Datei nicht gefunden: {path}"

        wb = load_workbook(path, read_only=True, data_only=not preserve_formulas)

        # Get worksheet
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                wb.close()
                return f"Fehler: Arbeitsblatt '{sheet_name}' nicht gefunden"
            ws = wb[sheet_name]
        else:
            ws = wb.active

        # Read range
        cells = ws[range_ref.upper()]

        # Convert to list of lists
        rows = []
        for row in cells:
            row_data = []
            for cell in row:
                value = cell.value
                row_data.append(value if value is not None else "")
            rows.append(row_data)

        wb.close()

        # Convert to dictionaries if include_header
        if include_header and len(rows) > 1:
            headers = rows[0]
            data_rows = []
            for row in rows[1:]:
                row_dict = {}
                for i, header in enumerate(headers):
                    if i < len(row):
                        row_dict[str(header)] = row[i]
                data_rows.append(row_dict)
            return json.dumps(data_rows, ensure_ascii=False, indent=2)
        else:
            return json.dumps(rows, ensure_ascii=False, indent=2)

    except Exception as e:
        return f"Fehler: {str(e)}"


@mcp.tool()
def excel_read_sheet(file_path: str, sheet_name: str = None, max_rows: int = 1000, preserve_formulas: bool = False) -> str:
    """
    Liest ein ganzes Arbeitsblatt und gibt es als JSON zurück.

    Erste Zeile wird als Spaltenüberschrift verwendet.

    Args:
        file_path: Pfad zur Excel-Datei (.xlsx)
        sheet_name: Name des Arbeitsblatts (optional, Standard: erstes Blatt)
        max_rows: Maximale Anzahl Zeilen (Standard: 1000, zum Schutz vor großen Dateien)
        preserve_formulas: Formeln als String lesen statt berechnete Werte (Standard: false)

    Returns:
        JSON-Array mit Zeilen als Dictionaries (Formeln als "=SUMME(...)" wenn preserve_formulas=true)
    """
    try:
        path = _resolve_path(file_path)
        if not path.exists():
            return f"Fehler: Datei nicht gefunden: {path}"

        wb = load_workbook(path, read_only=True, data_only=not preserve_formulas)

        # Get worksheet
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                wb.close()
                return f"Fehler: Arbeitsblatt '{sheet_name}' nicht gefunden"
            ws = wb[sheet_name]
        else:
            ws = wb.active

        # Get all rows
        rows = list(ws.iter_rows(values_only=True, max_row=max_rows + 1))

        wb.close()

        if len(rows) == 0:
            return json.dumps([], ensure_ascii=False)

        # First row as headers
        headers = [str(h) if h is not None else f"Column{i+1}" for i, h in enumerate(rows[0])]

        # Convert to dictionaries
        data_rows = []
        for row in rows[1:]:
            row_dict = {}
            for i, header in enumerate(headers):
                if i < len(row):
                    value = row[i]
                    row_dict[header] = value if value is not None else ""
            data_rows.append(row_dict)

        result = {
            "sheet": sheet_name or wb.sheetnames[0],
            "rows": len(data_rows),
            "columns": len(headers),
            "data": data_rows
        }

        return json.dumps(result, ensure_ascii=False, indent=2)

    except Exception as e:
        return f"Fehler: {str(e)}"


@mcp.tool()
def excel_write_cell(file_path: str, cell_ref: str, value: str, sheet_name: str = None) -> str:
    """
    Schreibt einen Wert in eine einzelne Zelle.

    Erstellt die Datei, falls sie nicht existiert.

    FORMELN: Strings die mit "=" beginnen werden automatisch als Formeln geschrieben.
    Beispiel: value="=SUM(A1:A10)" → Excel-Formel in Zelle

    Args:
        file_path: Pfad zur Excel-Datei (.xlsx)
        cell_ref: Zellenreferenz (z.B. 'A1', 'B5')
        value: Wert der geschrieben werden soll (oder Formel mit "=")
        sheet_name: Name des Arbeitsblatts (optional, Standard: erstes Blatt)

    Returns:
        Erfolgsmeldung oder Fehler
    """
    try:
        path = _resolve_path(file_path)

        # Load or create workbook
        if path.exists():
            wb = load_workbook(path)
        else:
            wb = Workbook()
            path.parent.mkdir(parents=True, exist_ok=True)

        # Get or create worksheet
        if sheet_name:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(sheet_name)
        else:
            ws = wb.active

        # Write cell
        ws[cell_ref.upper()] = value

        # Save
        wb.save(path)
        wb.close()

        return f"OK: Zelle {cell_ref} in '{sheet_name or 'Sheet1'}' geschrieben"

    except Exception as e:
        return f"Fehler: {str(e)}"


@mcp.tool()
def excel_write_range(file_path: str, range_ref: str, data: str, sheet_name: str = None) -> str:
    """
    Schreibt Daten in einen Zellenbereich.

    Erstellt die Datei, falls sie nicht existiert.

    FORMELN: Strings die mit "=" beginnen werden automatisch als Formeln geschrieben.
    Beispiel: [["Summe", "=SUM(A1:A10)"]] → Excel-Formel in Zelle

    Args:
        file_path: Pfad zur Excel-Datei (.xlsx)
        range_ref: Bereichsreferenz (z.B. 'A1:C10')
        data: JSON-Array mit Daten (2D-Array oder Array von Dictionaries)
        sheet_name: Name des Arbeitsblatts (optional, Standard: erstes Blatt)

    Returns:
        Erfolgsmeldung oder Fehler
    """
    try:
        path = _resolve_path(file_path)

        # Parse data
        try:
            data_parsed = json.loads(data) if isinstance(data, str) else data
        except json.JSONDecodeError as e:
            return f"Fehler: Ungültige JSON-Daten - {str(e)}"

        if not isinstance(data_parsed, list):
            return "Fehler: Daten müssen ein Array sein"

        # Load or create workbook
        if path.exists():
            wb = load_workbook(path)
        else:
            wb = Workbook()
            path.parent.mkdir(parents=True, exist_ok=True)

        # Get or create worksheet
        if sheet_name:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(sheet_name)
        else:
            ws = wb.active

        # Parse range
        (start_row, start_col), (end_row, end_col) = _parse_range(range_ref)

        # Convert dictionaries to 2D array if needed
        if data_parsed and isinstance(data_parsed[0], dict):
            headers = list(data_parsed[0].keys())
            rows = [[row.get(h, "") for h in headers] for row in data_parsed]
            data_parsed = [headers] + rows

        # Write data
        row_idx = start_row
        for row_data in data_parsed:
            if row_idx > end_row:
                break
            col_idx = start_col
            for value in row_data:
                if col_idx > end_col:
                    break
                ws.cell(row=row_idx, column=col_idx, value=value)
                col_idx += 1
            row_idx += 1

        # Save
        wb.save(path)
        wb.close()

        return f"OK: Bereich {range_ref} in '{sheet_name or 'Sheet1'}' geschrieben"

    except Exception as e:
        return f"Fehler: {str(e)}"


@mcp.tool()
def excel_write_sheet(file_path: str, data: str, sheet_name: str = None, clear_existing: bool = True) -> str:
    """
    Schreibt komplettes Arbeitsblatt mit neuen Daten (ohne Range-Angabe).

    Überschreibt das gesamte Arbeitsblatt. Automatische Größenerkennung.
    Ideal für Full-Roundtrip: excel_read_sheet() → Verarbeitung → excel_write_sheet()

    Erstellt die Datei/Arbeitsblatt, falls sie nicht existiert.

    FORMELN: Strings die mit "=" beginnen werden automatisch als Formeln geschrieben.
    Beispiel: {"Total": "=SUM(A1:A10)"} → Excel-Formel in Zelle

    Args:
        file_path: Pfad zur Excel-Datei (.xlsx)
        data: JSON-Array mit Daten (Array von Dictionaries oder 2D-Array)
        sheet_name: Name des Arbeitsblatts (optional, Standard: erstes Blatt)
        clear_existing: Bestehende Daten löschen (Standard: true)

    Returns:
        Erfolgsmeldung mit Anzahl geschriebener Zeilen
    """
    try:
        path = _resolve_path(file_path)

        # Parse data
        try:
            data_parsed = json.loads(data) if isinstance(data, str) else data
        except json.JSONDecodeError as e:
            return f"Fehler: Ungültige JSON-Daten - {str(e)}"

        if not isinstance(data_parsed, list):
            return "Fehler: Daten müssen ein Array sein"

        if len(data_parsed) == 0:
            return "Fehler: Daten-Array ist leer"

        # Load or create workbook
        if path.exists():
            wb = load_workbook(path)
        else:
            wb = Workbook()
            path.parent.mkdir(parents=True, exist_ok=True)

        # Get or create worksheet
        if sheet_name:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(sheet_name)
        else:
            ws = wb.active

        # Clear existing data if requested
        if clear_existing:
            # Delete all rows (keep worksheet)
            ws.delete_rows(1, ws.max_row)

        # Convert dictionaries to 2D array if needed
        if isinstance(data_parsed[0], dict):
            headers = list(data_parsed[0].keys())
            rows_data = [[row.get(h, "") for h in headers] for row in data_parsed]
            all_rows = [headers] + rows_data
        else:
            all_rows = data_parsed

        # Write all data starting from A1
        for row_idx, row_data in enumerate(all_rows, start=1):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Save
        wb.save(path)
        wb.close()

        num_data_rows = len(all_rows) - 1 if isinstance(data_parsed[0], dict) else len(all_rows)
        num_cols = len(all_rows[0]) if all_rows else 0
        sheet_display_name = sheet_name or "Sheet1"

        return f"OK: {num_data_rows} Zeilen × {num_cols} Spalten in '{sheet_display_name}' geschrieben"

    except Exception as e:
        return f"Fehler: {str(e)}"


@mcp.tool()
def excel_create_sheet(file_path: str, sheet_name: str) -> str:
    """
    Erstellt ein neues Arbeitsblatt.

    Erstellt die Datei, falls sie nicht existiert.

    Args:
        file_path: Pfad zur Excel-Datei (.xlsx)
        sheet_name: Name des neuen Arbeitsblatts

    Returns:
        Erfolgsmeldung oder Fehler
    """
    try:
        path = _resolve_path(file_path)

        # Load or create workbook
        if path.exists():
            wb = load_workbook(path)
        else:
            wb = Workbook()
            path.parent.mkdir(parents=True, exist_ok=True)

        # Check if sheet already exists
        if sheet_name in wb.sheetnames:
            wb.close()
            return f"Fehler: Arbeitsblatt '{sheet_name}' existiert bereits"

        # Create sheet
        wb.create_sheet(sheet_name)

        # Save
        wb.save(path)
        wb.close()

        return f"OK: Arbeitsblatt '{sheet_name}' erstellt"

    except Exception as e:
        return f"Fehler: {str(e)}"


@mcp.tool()
def excel_delete_sheet(file_path: str, sheet_name: str) -> str:
    """
    Löscht ein Arbeitsblatt.

    Args:
        file_path: Pfad zur Excel-Datei (.xlsx)
        sheet_name: Name des zu löschenden Arbeitsblatts

    Returns:
        Erfolgsmeldung oder Fehler
    """
    try:
        path = _resolve_path(file_path)

        if not path.exists():
            return f"Fehler: Datei nicht gefunden: {path}"

        wb = load_workbook(path)

        # Check if sheet exists
        if sheet_name not in wb.sheetnames:
            wb.close()
            return f"Fehler: Arbeitsblatt '{sheet_name}' nicht gefunden"

        # Don't allow deleting the last sheet
        if len(wb.sheetnames) == 1:
            wb.close()
            return "Fehler: Kann das letzte Arbeitsblatt nicht löschen"

        # Delete sheet
        del wb[sheet_name]

        # Save
        wb.save(path)
        wb.close()

        return f"OK: Arbeitsblatt '{sheet_name}' gelöscht"

    except Exception as e:
        return f"Fehler: {str(e)}"


if __name__ == "__main__":
    mcp.run()
